# -*- coding: utf-8 -*-
import os, csv, re, glob
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

BASE = r"D:\Iteracion\sheets_glmex2"
DAILY = os.path.join(BASE, "daily")
OUT = os.path.join(BASE, "GL-MEX2_workbook.xlsx")

INT_RE = re.compile(r"^-?\d+$")
FLT_RE = re.compile(r"^-?\d+\.\d+$")

HEADER_FILLS = PatternFill("solid", fgColor="1F4E78")
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
TOTAL_FILL = PatternFill("solid", fgColor="FCE4D6")

def cast(v):
    if INT_RE.match(v): return int(v)
    if FLT_RE.match(v): return float(v)
    return v

def add_sheet(wb, name, rows):
    ws = wb.create_sheet(title=name[:31])
    for r in rows:
        ws.append([cast(c) for c in r])
    # styling
    for i, r in enumerate(rows, start=1):
        first = r[0] if r else ""
        if i == 1:  # title
            ws.cell(i,1).font = Font(bold=True, size=13, color="1F4E78")
        elif first in ("DETALLE POR VIDEO (META)", "META vs SHOPIFY (POR PRODUCTO)", "CONCLUSIONES"):
            c = ws.cell(i,1); c.font = Font(bold=True, color="FFFFFF"); c.fill = HEADER_FILLS
        elif first in ("Producto",):  # column header rows
            for j in range(1, len(r)+1):
                c = ws.cell(i,j); c.font = Font(bold=True); c.fill = SECTION_FILL
        elif first == "TOTAL":
            for j in range(1, len(r)+1):
                c = ws.cell(i,j); c.font = Font(bold=True); c.fill = TOTAL_FILL
    # widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 16
    for col in "CDEFGHIJKLMNOPQR":
        ws.column_dimensions[col].width = 12
    ws.freeze_panes = "A7"
    return ws

def read_csv(path):
    with open(path, "r", encoding="utf-8-sig") as f:
        return [row for row in csv.reader(f)]

wb = Workbook()
wb.remove(wb.active)

# 1) Mes  2) 7 dias  3) daily (newest first)
add_sheet(wb, "Mes (MTD)", read_csv(os.path.join(BASE,"mes.csv")))
add_sheet(wb, "7 dias", read_csv(os.path.join(BASE,"7dias.csv")))

daily_files = sorted(glob.glob(os.path.join(DAILY, "Dia_*.csv")), reverse=True)
for path in daily_files:
    date = os.path.basename(path).replace("Dia_","").replace(".csv","")
    add_sheet(wb, f"Dia {date}", read_csv(path))

wb.save(OUT)
print("XLSX generado:", OUT)
print("Pestanas:", wb.sheetnames)
